# coding=utf-8

import urllib.request

from openpyxl import Workbook
from openpyxl import load_workbook
import time
import datetime
import threading
from operator import itemgetter, attrgetter 
import os
import pickle
import math

#for progress
g_index = 0

g_mutex = threading.Lock()
g_allDataList = []

#tuple (code, name, price, pe, pb, roa, roe, roaIndex, roeIndex, overallIndex)
class StockData():
	#roa: Return on Assets,  收益率
	#roe: Return on Equity, 投资回报率
	def __init__(self, code, name, price, pe, pb, roa, roe, roaIndex = -1, roeIndex = -1, overallIndex = -1):
		#股票代码:{0} 股票名称:{1} 最新价:{2} 市盈率:{3} 市净率:{4} 收益率:{5} 回报率:{6}
		self.code = code
		self.name = name
		self.price = price
		self.boll = 100.0
		self.ene = 100.0
		self.pe = pe
		self.pb = pb
		self.roa = roa
		self.roe = roe
		self.roaIndex = roaIndex
		self.roeIndex = roeIndex
		self.overallIndex = overallIndex
		
		self.array = []
		
	def __repr__(self):
		return repr((self.code, self.name, self.price,self.boll, self.ene, self.pe,self.pb,self.roa,self.roe,self.roaIndex,self.roeIndex,self.overallIndex, self.array))


		
class MagicFormula():

	def __init__(self):

		#time stamp
		now = datetime.datetime.now()
		self.__timeStamp = now.strftime("%Y_%m_%d_%H_%M_%S")
		
		#IO
		self.__topCount = 50
		
		self.__excel_filepath = 'E:/SmartMoney/quantization/test/Data_%s.xlsx' % self.__timeStamp
		self.__txt_filepath = 'E:/SmartMoney/quantization/test/Data_%s.txt' % self.__timeStamp
		
		self.__boll_txt_filepath = 'E:/SmartMoney/quantization/test/BollData_%s.txt' % self.__timeStamp
		
		#debug info
		self.__quickTest = False
		self.__printDetails = False
		self.__printThreadingInfo = False
		self.__printHeadsupInfo = True
		self.__printProgress = True
		self.__printTimeInfo = True
		
		#data list
		self.__finalSortedList = []
		
		#database URL to query
		self.__URL = 'http://qt.gtimg.cn/q='
		
		#parameters for BOLL
		self.__k = 2.0
		self.__n = 20
		
		#parameters for ENE
		self.__N_ENE = 10
		self.__M1_ENE = 11
		self.__M2_ENE = 9
		
		#N:=11;M1:=10;M2:=9;


	def main(self):
		#start timer
		startTime = time.time()
		
		if self.__printHeadsupInfo :
			print(">>开始:收集股票数据")
		
		#create all threads
		self.CreateThreadPool()
		
		#start all threads
		for thread in self.__thread_list:
			thread.start()
	 
		# 主线程中等待所有子线程退出
		for thread in self.__thread_list:
			thread.join()
			
		if self.__printHeadsupInfo :
			print("<<结束:收集股票数据")

		#sorting data
		self.SortingDataList()
		
		#write to Excel file
		self.WriteToExcel()
		
		#write to txt file
		self.WriteToTxt()
		
		#end the timer
		endTime = time.time()
		if self.__printTimeInfo :
			print ('***all time: ' + str(endTime-startTime))
		

	def CreateThreadPool(self):
		self.__thread_list = list()
		# 先创建线程对象
		#multiple threads to collect data at the same time
		#沪市股票 code in range(600000, 604000)
		#深市主板股票 code in range(0, 1999) = #[0,1000]  + 1696 + 1896
		#深市中小板股票 code in range(2000, 2999)
		#创业板股票 code in range(300000, 300500)
		steps = 200
		threadCount = 0
		if not self.__quickTest :
			#沪市股票
			self.PutToThreads(600000, 604000, steps, threadCount)
			threadCount += (604000-600000)/steps
			
			#深市主板股票
			self.PutToThreads(0, 1000, steps, threadCount)
			threadCount += (1000-0)/steps
			
			#深市中小板股票
			self.PutToThreads(2000, 2800, steps, threadCount)
			threadCount += (2800-2000)/steps
			
			#创业板股票
			steps = 250
			self.PutToThreads(300000, 300500, steps, threadCount)
			threadCount += (300500-300000)/steps
			
			thread_name = "thread_" + str(threadCount)
			self.__thread_list.append(threading.Thread(target = self.GetDataSets, name = thread_name, args = (1695,1700)))
			threadCount +=1
			thread_name = "thread_" + str(threadCount)
			self.__thread_list.append(threading.Thread(target = self.GetDataSets, name = thread_name, args = (1895,1900)))
		else:
			#quick test
			steps = 5
			self.PutToThreads(600000, 600010, steps, threadCount)


	def PutToThreads(self, startCode, endCode, steps, existingThreadCount):
		thread_num = (endCode-startCode)/steps
		for i in range(0, int(thread_num)):
			threadIndex = existingThreadCount + i
			thread_name = "thread_%s" % threadIndex
			self.__thread_list.append(threading.Thread(target = self.GetDataSets, name = thread_name, args = (startCode+i*steps,startCode+(i+1)*steps)))
		
		existingThreadCount += thread_num
		

	def GetDataSets(self, startCode, endCode):
		#try:
		for code in range(int(startCode), int(endCode)):
			status = self.GetCodeData(code)

	def GetCodeData(self, code):
		url = self.GetURL(code)
		if len(url) < 1 :
			return False
		
		req = urllib.request.Request(url)
		content = urllib.request.urlopen(req).read()
		str = content.decode('gbk')
		strList = str.split('"')
		if len(strList) < 2 :
			return False
		data = strList[1].split('~')
		if len(data) < 47 :
			return False

		name = "%-6s" % data[1]
		stockCode = "%-6s" % data[2]

		if len(name) < 1 or len(stockCode) < 1 :
			return False

		currentPrice = -1.0
		PE = -1.0
		PB = -1.0
		ROE = -1.0
		ROA = -1.0
		
		if len(data[3]) > 0 :  
			currentPrice = float(data[3])  
			
		if len(data[39]) > 0 :
			PE = float(data[39])
   
		if len(data[46]) > 0 :
			PB = float(data[46])

		#ROE = PB/PE
		if PE > 0.0000001 :
			ROE = 100.0*PB/PE
		#1/PE
		#ROA: Return on Assets
		if PE > 0.0000001 :
			ROA = 100.0/PE
			
		if self.__printDetails :
			print("股票代码:{0} 股票名称:{1} 最新价:{2} 市盈率:{3} 市净率:{4} 收益率:{5} 回报率:{6}".format(stockCode, name, currentPrice, PE, PB, ROA, ROE) )
			#print("股票名称:{0} 收益率:{1} 回报率:{2} ".format(name, ROA, ROE ))
		
		#(code, name, price, pe, pb, roa, roe, roaIndex, roeIndex, overallIndex)
		stockData = StockData(stockCode, name, currentPrice, PE, PB, ROA, ROE, -1, -1, -1)
		self.GetBollAndEneData(code, stockData)
		
		
		global g_index, g_mutex, g_allDataList
		if g_mutex.acquire():
			g_allDataList.append(stockData)
			if self.__printProgress :
				print (g_index)
				
			g_index += 1
			g_mutex.release()

		return True


	def GetURL(self, code):
		url = ''
		codeName = ''
		
		if code >= 600000 and code <=604000 :
			#沪市股票 code in range(600000, 602100)
			# http://qt.gtimg.cn/q=sh600001
			codeName = 'sh' + str(code)
		elif code >= 0 and code <=2999 :
			#深市主板股票 code in range(0, 1999) = #[0,1000]  + 1696 + 1896
			#深市主板股票 code in range(0, 1999)
			#深市中小板股票 code in range(2000, 2999)
			# http://qt.gtimg.cn/q=sz000858
			codeName = "sz%06d" % code
		elif code >= 300000 and code <=300500 :
			#创业板股票 code in range(300000, 300500)
			# http://qt.gtimg.cn/q=sz300001
			codeName = 'sz' + str(code)
		else :
			codeName = ''
			
		if self.__printThreadingInfo :
			print ("%s :  I come from %s" % (codeName, threading.currentThread().getName()))
			
		if len(codeName) > 0 :
			url = self.__URL + codeName
		
		else:
			url = ''
		
		return url


	def SortingDataList(self):
		global g_allDataList

		#sorting by roa
		list_sortedViaROA = sorted(g_allDataList, key=attrgetter('roa'), reverse = True)
		
		#update roaIndex
		i = 0
		for item in list_sortedViaROA :
			list_sortedViaROA[i].roaIndex = i
			i+=1
		
		#sorting by roe
		list_sortedViaROE = sorted(g_allDataList, key=attrgetter('roe'), reverse = True)
		
		#update roeIndex and overallIndex
		i = 0
		for item in list_sortedViaROE :
			list_sortedViaROE[i].roeIndex = i
			list_sortedViaROE[i].overallIndex = i + list_sortedViaROE[i].roaIndex
			i+=1

		#sorting by overallIndex (mix roaIndex and roeIndex)
		self.__finalSortedList = sorted(g_allDataList, key=attrgetter('overallIndex'))

		
	#use openpyxl
	def WriteToExcel(self):
		workbook = Workbook()
		worksheet = workbook.create_sheet('All Data', 0)

		#Make Excel Top Row
		worksheet['A1'] = "股票代码"
		worksheet['B1'] = "股票名称"
		worksheet['C1'] = "最新价"
		worksheet['D1'] = "BOLL"
		worksheet['E1'] = "ENE"
		worksheet['F1'] = "市盈率"
		worksheet['G1'] = "市净率"
		worksheet['H1'] = "收益率ROA(%)"
		worksheet['I1'] = "回报率ROE(%)"
		worksheet['J1'] = "收益率index"
		worksheet['K1'] = "回报率index"
		worksheet['L1'] = "index+index"

		i = 2
		for item in  self.__finalSortedList:
			#code, name, price, pe, pb, roa, roe, roaIndex = -1, roeIndex = -1, overallIndex
			cell = worksheet.cell(row = i, column = 1).value =  item.code
			cell = worksheet.cell(row = i, column = 2).value =  item.name
			cell = worksheet.cell(row = i, column = 3).value =  item.price
			cell = worksheet.cell(row = i, column = 4).value =  item.boll
			cell = worksheet.cell(row = i, column = 5).value =  item.ene
			cell = worksheet.cell(row = i, column = 6).value =  item.pe
			cell = worksheet.cell(row = i, column = 7).value =  item.pb
			cell = worksheet.cell(row = i, column = 8).value =  item.roa
			cell = worksheet.cell(row = i, column = 9).value =  item.roe
			cell = worksheet.cell(row = i, column = 10).value =  item.roaIndex
			cell = worksheet.cell(row = i, column = 11).value =  item.roeIndex
			cell = worksheet.cell(row = i, column = 12).value = item.overallIndex
			i +=1

		workbook.save(self.__excel_filepath)
		
		if self.__printHeadsupInfo :
			print("<<Done! Excel file saved.")
	
	def WriteToTxt(self):
		#open file
		fileHandle = open(self.__txt_filepath,'w+')
		
		#Write top line
		#股票代码:{0} 股票名称:{1} 最新价:{2} 市盈率:{3} 市净率:{4} 收益率:{5} 回报率:{6}
		fileHandle.write(
				#self.__timeStamp + ' \n' +
				#'0    ' +
				'股票代码    ' +
				'股票名称    ' + 
				'最新价    ' + 
				'BOLL    ' + 
				'ENE    ' + 
				'市盈率    ' +
				'市净率    ' + 
				'收益率ROA(%)  ' +
				'回报率ROE(%)  ' +
				'收益率index  ' +
				'回报率index  ' +
				'index+index  \n'
				) 
		i = 1
		for item in  self.__finalSortedList:
			#code, name, price, boll, ene, pe, pb, roa, roe, roaIndex = -1, roeIndex = -1, overallIndex
			fileHandle.write(
							#str(i) + '    ' +
							str(item.code) + '    ' +
							str(item.name)+ '  ' + 
							str(item.price)+ '    ' + 
							str(item.boll)+ '    ' +
							str(item.ene)+ '    ' +
							str(item.pe)+ '    ' +
							str(item.pb)+ '    ' + 
							str(item.roa)+ '    ' +
							str(item.roe)+ '    ' +
							str(item.roaIndex)+ '    ' +
							str(item.roeIndex)+ '    ' +
							str(item.overallIndex) + '\n'
							)
			#if i is (self.__topCount-1) :
			#	fileHandle.write('***********The above is TOP 50 *************\n\n')
			
			i += 1

		#close file
		fileHandle.close()
		
		if self.__printHeadsupInfo :
			print("<<Done! txt file saved.")


	def ComputeMA(self, n, stockData):
		MA = 0.0
		array = stockData.array
		if len(array) < n :
			print (array)
			print ('error: history data was less than %s days' % n)
			return MA
		
		for i in range(n):
			MA += array[i]
			
		MA = MA/float(n)
		return MA
			
	#compute ENE
	#N:=10;M1:=11;M2:=9;
	#UPPER:(1+M1/100)*MA(CLOSE,N);
	#LOWER:(1-M2/100)*MA(CLOSE,N);
	#ENE:(UPPER+LOWER)/2;
	def ComputeENE(self, stockData):
		ene = 100.0
		array = stockData.array
		MA = 0.0
		MA = self.ComputeMA(self.__N_ENE, stockData)
		UPPER = (1.0+self.__M1_ENE/100.0)*MA
		LOWER = (1.0-self.__M2_ENE/100.0)*MA
		ENE = (UPPER+LOWER)/2.0;
		#print (stockData.name)
		#print ('upper %s' % UPPER)
		#print ('lower %s' % LOWER)
		
		if (UPPER - LOWER) <= 0.0 or array[0] == 0.0 :
			stockData.ene = ene
			return ene
			
		percentENE = (array[0] - LOWER)/(UPPER - LOWER)
		stockData.ene = percentENE
		return percentENE
			
	#compute Boll
	def ComputeBoll(self, stockData):
		k = self.__k
		n = self.__n
		array = stockData.array
		boll = 100.0
		if len(array) < n :
			print (array)
			print ('error: history data was less than %s days' % n)
			stockData.boll = boll
			return boll
		
		MA = 0.0
		MA = self.ComputeMA(n, stockData)
		
		MD = 0.0
		for i in range(n):
			MD += (array[i]-MA)*(array[i]-MA)
		
		MD = MD/float(n)
		MD = math.sqrt(MD)
		
		#MB= 0.0
		#for i in range(n):
		#	MB += array[i]
			
		#MB = MB - array[0]
		#MB = MB/float(n-1)
		MB = MA
		UP = MB + k*MD
		DN = MB - k*MD
		#print ('down = %s' % DN)
		#print ('up = %s' % UP)
		#print ('middle = %s' % MB)
		#print ('price = %s' % array[0])
		if (UP - DN) <= 0.0 or array[0] == 0.0 :
			stockData.boll = boll
			return boll
			
		percentB = (array[0] - DN)/(UP - DN)
		stockData.boll = percentB
		#print ("b percent= %s" % percentB)
			
		return percentB
			
	def GetBollAndEneData(self, code, stockData):
		
		codeString = str(code)
		now = datetime.datetime.now()
		nowString = now.strftime("%Y%m%d")
		thisYear = now.year
		startTime = now.replace(year = thisYear-1)
		starttimeString = startTime.strftime("%Y%m%d")
		
		#http://quotes.money.163.com/service/chddata.html?code=0600000&start=20000720&end=20151222
		
		url = 'http://quotes.money.163.com/service/chddata.html?code='
		if code >= 600000 :
			url += '0'
		else:
			url += '1'
			codeString = "%06d" % code
			

		url += codeString
		url += '&start='
		url +=  starttimeString
		url += '&end='
		url += nowString
		
		req = urllib.request.Request(url)
		content = urllib.request.urlopen(req).read()
		stri = content.decode('gbk')
		#print (stri)
		strList = stri.split('\n')
		if len(strList) < 2 :
			return False

		i = 0
		for item in strList:
			itemWithoutSpace = item.strip()

			if len(itemWithoutSpace) > 0:
				data = itemWithoutSpace.split(',')
				if i > 0:
					stockData.array.append(float(data[3]))
					
			i +=1
			#i > 0 and i < (self.__n + 5) and 	

		#boll = self.ComputeBoll(stockData)
		stockData.boll = self.ComputeBoll(stockData)
		stockData.ene = self.ComputeENE(stockData)
		
		print (stockData.boll)
		print (stockData.ene)


if __name__ == '__main__':
    MagicFormula().main()


